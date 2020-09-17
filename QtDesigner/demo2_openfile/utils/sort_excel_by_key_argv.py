# encoding=utf-8
from sys import argv

import pandas as pd
import os
import operator
from pandas import DataFrame
import pprint
from openpyxl.utils import get_column_letter

"""
 py文件相同目录下要有yaml_read.py文件，还要有个子文件夹data_source存放key_words.yml文件
 把excel表格中沙河税务所的行筛查出来，再按照尾号排序，另存一份
 步骤：
 1.读取文件
 2.按照事先指定的关键字列表，确定两个关键列：主管税务所、纳税人识别码 ,注意要先把空格过滤掉 
     search_keyword(object_keyword, keywords_list) -> final_keyword
 3.提取出沙河所的列 pick_data(pick_list, key, value)
 4.按照尾号排序 sort_data(sort_list, by_key)
 5.保存文件   
"""


def search_keyword(origin_list, keywords_list):
    """
     寻找对应的表头，如果表格中的标题列与关键词列有相同的字符串，
     那么就把这个列当做要操作的列，可以进行排序，筛选等
    :param origin_list: 表格中的标题列
    :param keywords_list:  事先写好的关键词列
    :return:
    """
    header_list = [header for header in origin_list[0]]
    for header_k in header_list:
        for k in keywords_list:
            if header_k == k:
                return k


def pick_data(pick_list, key, value):
    """
    pick data from dicts in a list
    从列表中的字典中寻找指定的内容，并返回符合条件的字典存入字典中
    :param pick_list: 需要进行筛选的原始列表
    :param key: dict's key
    :param value: dict's value
    :return: list after picking
    """
    list_picked = []
    for row_dict in pick_list:
        # row_dict[key] 有可能是数字类型的，所以要转成str才能进行查询
        if value in str(row_dict[key]):
            list_picked.append(row_dict)
    return list_picked


def find_the_one_digital_in_str(origin_str, number=-1):
    """
     find_the_digital_from_str
     根据指定的字符串和number，返回第number位的数字，从0开始，倒数第一位是-1(默认值)
    :param origin_str:
    :param number: which number of the digital do you want, the first is zero, the last is -1(default)
    :return: one digital in the str you want, if no digital in the string, then return 0
    """
    list_digitals = [i for i in str(origin_str) if i.isdigit()]
    # 如果当前cell中没有数字，则返回''
    if list_digitals:
        the_digital = list_digitals[number]
        return the_digital
    return '0'


def sort_data(sort_list, by_key, order='ascend'):
    """
    按照关键字排序
    :param sort_list: 要排序的list，里边存放的是每一行的字典数据（row_dict)
    :param by_key: 被排序的列的标题名
    :param order:
    :return: 排好序的列
    """
    list_sorted = []
    # list_sorted = sorted(sort_list, key=operator.itemgetter(by_key))
    for i in range(10):
        # 从0到9遍历数字，按尾号排序
        for row_dict in sort_list:
            cell_value = row_dict[by_key]
            # 如果
            if find_the_one_digital_in_str(cell_value) is None:
                list_sorted.append(row_dict)
                continue

            if find_the_one_digital_in_str(cell_value) == str(i):
                list_sorted.append(row_dict)
    return list_sorted;


def sort_file(filename, path='', keyword_department='', list_master_keywords=[], list_taxpayer_number_keywords=[]):
    print('sort_file, file name: ', filename)

    suffix_filename = os.path.splitext(filename)[1]
    if suffix_filename == '.xlsx' or suffix_filename == '.xls':
        if keyword_department == '':
            # keyword_department = yaml_read.read_keywords_from_yaml()['keyword_department']
            keyword_department = '沙河'
        if not list_taxpayer_number_keywords:
            # list_taxpayer_number_keywords = yaml_read.read_keywords_from_yaml()['taxpayer_number_keywords']
            list_taxpayer_number_keywords = ['纳税人识别号', '纳税识别号', '纳税人识别码',
                                             '纳税识别码', '社会信用代码', '社会信用代码(纳税人识别号)',
                                             '社会信用代码(纳税识别号)', '社会信用代码（纳税人识别号）',
                                             '社会信用代码（纳税识别号）', '税号', '信用代码', '营业执照号码',
                                             '统一社会信用代码', '纳税人税号', '企业识别号', '企业税号', 'SHXYDM']
        if not list_master_keywords:
            # list_master_keywords = yaml_read.read_keywords_from_yaml()['master_keywords']
            list_master_keywords = ['主管所', '主管税务所', '税务所', '管理所', '外围所',
                                    '主管税务所（科、分局）', '主管税务所(科、分局)']
        # read excel file to list(dicts)
        try:
            xl = pd.ExcelFile(path + filename)
        except:
            message = '找不到此文件，请重试'
            return '', message
        sheet_names = xl.sheet_names  # get all name of sheets
        print('sheet_names: ', sheet_names)
        # 保存做过处理的sheet名
        sorted_sheets = []

        # set width of column
        WIDTH = 19

        # 遍历每一个sheet
        for sheet_num_int, sheet_name_str in enumerate(sheet_names):
            print("begin sheet: " + sheet_name_str)
            df = pd.read_excel(path + filename, sheet_name=sheet_name_str)
            if df.keys().size == 0:
                print(sheet_name_str + ' is empty.')
                continue
            data_list = df.to_dict(orient='records')
            # print(df.to_dict(orient='splite')['columns'])
            # 取得原表的表头，输出时候还按照原表头的顺序
            order = df.to_dict(orient='splite')['columns']
            # title_dict = data_list[0]
            # order = list(title_dict.keys())

            keyword_master = search_keyword(data_list, list_master_keywords)
            # print('keyword_master', keyword_master)
            keyword_taxpayer = search_keyword(data_list, list_taxpayer_number_keywords)
            # print('keyword_taxpayer', keyword_taxpayer)

            # 如果标题栏中找不到纳税人信息，则不进行数据清洗,原样输出
            new_list = data_list
            if keyword_taxpayer is None:
                print('can not find taxpayers keywords in sheet: ' + sheet_name_str)
            else:
                if keyword_master is None:
                    # 如果标题栏没有部门信息，就要筛选出本部门的内容
                    pass
                else:
                    # 如果标题栏含有部门信息，就要筛选出本部门的内容
                    new_list = pick_data(data_list, keyword_master, keyword_department)
                    if new_list == []:
                        continue

                # 无论有没有部门信息，都会按纳税人识别号尾号排序
                sorted_sheets.append(sheet_name_str)
                new_list = sort_data(new_list, keyword_taxpayer)
                # print('排序后的结果:', new_list)

            # write list into a new excel file
            df = pd.DataFrame(new_list)
            df = df[order]
            prefix_filename = os.path.splitext(filename)[0]
            suffix_filename = os.path.splitext(filename)[1]
            # new_filename = prefix_filename + '_已排序' + suffix_filename
            new_filename = prefix_filename + '_已排序.xlsx'
            print('sheet_num_int:', sheet_num_int)
            try:
                if sheet_num_int == 0:
                    writer = pd.ExcelWriter(path + new_filename)
                    df.to_excel(writer, sheet_name=sheet_name_str, index=False)
                    work_sheet = writer.sheets[sheet_name_str]
                    # set width of all columns
                    for i in range(1, work_sheet.max_column + 1):
                        work_sheet.column_dimensions[get_column_letter(i)].width = WIDTH
                    writer.save()
                    writer.close()
                else:
                    with pd.ExcelWriter(path + new_filename, mode='a', engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name_str, index=False)
                        work_sheet = writer.sheets[sheet_name_str]
                        # set width of all columns
                        for i in range(1, work_sheet.max_column + 1):
                            work_sheet.column_dimensions[get_column_letter(i)].width = WIDTH
                print("end sheet: " + sheet_name_str)
            except Exception as e:
                message = '文件写入错误\n请尝试关闭此文件：%s' % new_filename
                return '', message
                print(e)

        if sorted_sheets:
            message = '筛选成功\n文件名是：' + new_filename
            return new_filename, message
        else:
            # 把已经新建的文件（原样输出）删除
            os.remove(new_filename)
            message = '没有找到可以排序或者筛选的数据，请核实原文件内容'
            return '', message
    else:
        message = '文件格式错误，只能处理EXCEL（以.xlsx 或.xls结尾）文件'
        return '', message
    '''
     os.remove((path+filename))
     os.rename(path+new_filename, path+filename)
     break # 暂时循环一次，只处理第一个sheet
     后续还有问题要解决：
       1.空sheet: if df.keys().size == 0: 已解决
       2.多个sheets的话如果用to_excel的话，后边的sheet会覆盖前边的。如果用ExcelWriter可以追加sheet在最后
           但是会在最左侧增加一条索引列，另外用新文件名的话会提示找不到，用原文件名会每次都追加一堆sheet
           已解决： 方法，处理第一个sheet时直接用to_excel写入一个新文件,后边的就先用with打开这个新文件
       3. 如果中间有空列或者没有数字的情况，遇到信用代码找尾号时就会报错
          已解决，如果遇到此情况，按照尾号零处理（排到最前边）或者直接略过（按尾号是空）
       4. 如果所有sheet都找不到要排序的关键字，需要给出提示 已解决
       5. 如果遇到表格不整齐，比如有合并格的处理 已解决：经测试目前无问题，可以忽略该列
       6. 给出统计数字，共计多少列，每个尾号多少列 未解决：对于多个sheet的表格不太实用
       7. 遇到错误的处理，比如文件已被其他软件占用（打开） 已解决
       8. 可以处理xlsx,xls格式的文件 
       9. 加入找不到文件的处理
       A. 更改列宽 已完成 每个列宽都相等，用WIDTH保存列宽变量
       B. 有税务所列但是没有关键字税务所，会发生错误 
          已解决：如果有部门列但是没有此部门数据，直接跳过这个sheet的操作，已排序的表里没有这个sheet
    '''


if __name__ == '__main__':
    # import yaml_read

    path = os.getcwd() + '\\'
    filename = "panda_02_read_excel_02.xlsx"
    filename = "7月未申报0716.xlsx"
    # filename = argv[1]
    # yaml_filename = 'key_words.yml'
    # with open(path+filename, 'r', encoding='utf-8') as f:
    #     s = f.read()
    # key_words = yaml.load(s, Loader=yaml.FullLoader)

    # list_taxpayer_number_keywords = yaml_read.read_keywords_from_yaml()['taxpayer_number_keywords']
    # list_master_keywords = yaml_read.read_keywords_from_yaml()['master_keywords']
    # keyword_department = yaml_read.read_keywords_from_yaml()['keyword_department']
    filename_str, message = sort_file(filename, path)
    # filename_str, message = sort_file(filename, path, keyword_department, list_master_keywords, list_taxpayer_number_keywords)
    try:
        # filename_str, message = sort_file(filename, path, keyword_department, list_master_keywords, list_taxpayer_number_keywords)
        print('filename', filename)
        print('message', message)
        pass
    except Exception as e:
        print(e)
    else:
        print('succeed')
